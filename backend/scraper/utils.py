import datetime
import os
from typing import Callable

from openpyxl import Workbook, load_workbook
from playwright.async_api import Page

from companies import FieldMapping, PaginationConfig, SearchConfig
import logging

log = logging.getLogger("job_scraper")


def _normalize_job(raw: dict, fm: FieldMapping, company: str) -> dict:
    """
    Convert a raw API payload dict → clean job dict using FieldMapping.
    Handles a few common special cases (location as dict/list, department as list).
    """
    job = {"company": company}

    job["title"] = _get(raw, fm.title)

    # location: may be a string, a dict {"name": ...}, or a list of dicts
    loc_raw = _get(raw, fm.location)
    if isinstance(loc_raw, list) and loc_raw:
        first = loc_raw[0]
        job["location"] = first.get("name") if isinstance(first, dict) else first
    elif isinstance(loc_raw, dict):
        job["location"] = loc_raw.get("name")
    else:
        job["location"] = loc_raw

    if company == "Uber":
        job["location"] = _extract_uber_location(raw)
    job["id"] = _get(raw, fm.job_id)

    # apply_url
    job["apply_url"] = _get(raw, fm.apply_url)
    if company == "Uber" and not job.get("apply_url") and job.get("id"):  # Uber API doesn't return apply URL's
        job["apply_url"] = f"https://www.uber.com/global/en/careers/list/{job['id']}/"

    # department: may be a dict or a list of dicts
    dept_values = []

    for key in fm.department:
        val = raw.get(key)

        if isinstance(val, list):
            for v in val:
                if isinstance(v, dict):
                    dept_values.append(v.get("name"))
                else:
                    dept_values.append(v)

        elif isinstance(val, dict):
            dept_values.append(val.get("name"))

        elif val:
            dept_values.append(val)

    # clean
    dept_values = [v for v in dept_values if v]
    dept_values = list(dict.fromkeys(dept_values))

    job["department"] = ", ".join(dept_values) if dept_values else None
    return {k: v for k, v in job.items() if v is not None}


def _get(d: dict, keys: list[str]):
    """Try each key in order, return first non-None value."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return None


def flatten_jobs(all_results):
    jobs = []
    for company, js in all_results.items():
        jobs.extend(js)
    return jobs


def deduplicate_jobs(jobs):
    seen = set()
    unique = []

    for job in jobs:
        key = job.get("id") or job.get("apply_url")

        if not key:
            continue

        if key in seen:
            continue

        if job.get("company") == "Rippling" and job.get("department") != "Engineering":
            continue

        seen.add(key)
        unique.append(job)

    return unique


def score_job(job):
    score = 0
    title = (job.get("title") or "").lower()
    location = (job.get("location") or "").lower()

    # role relevance
    if "software" in title: score += 5
    if "engineer" in title: score += 5
    if "backend" in title: score += 3
    if "frontend" in title: score += 3
    if "fullstack" in title: score += 3
    if "intern" in title: score += 2

    # location preference (customize)
    if "india" in location: score += 3
    if "remote" in location: score += 2

    if "senior" in title: score -= 10
    if "sr" in title: score -= 10

    return score


def sort_jobs(jobs):
    return sorted(jobs, key=lambda x: x.get("score", 0), reverse=True)


def add_scores(jobs):
    for job in jobs:
        job["score"] = score_job(job)
    return jobs


def save_final_excel(jobs):
    os.makedirs("runs", exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"runs/jobs_final_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Jobs"

    headers = ["Company", "Title", "Location", "Department", "Score", "ID", "Apply URL"]
    ws.append(headers)

    for job in jobs:
        if not job.get("title"):
            continue

        if job.get("score") < 5:
            continue
        ws.append([
            _safe_excel_value(job.get("company")),
            _safe_excel_value(job.get("title")),
            _safe_excel_value(job.get("location")),
            _safe_excel_value(job.get("department")),
            _safe_excel_value(job.get("score")),
            _safe_excel_value(job.get("id")),
            _safe_excel_value(job.get("apply_url")),
        ])

    # nice UX
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    log.info(f"\n📊 Saved FINAL Excel: {filename}")
    return filename


def save_to_excel(all_results: dict[str, list[dict]]):
    os.makedirs("runs", exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"runs/jobs_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # ✅ Header
    headers = ["Company", "Title", "Location", "Department", "ID", "Apply URL"]
    ws.append(headers)

    # ✅ Data
    for company, jobs in all_results.items():
        for job in jobs:
            if not job.get("title"):
                continue
            ws.append([
                _safe_excel_value(job.get("company")),
                _safe_excel_value(job.get("title")),
                _safe_excel_value(job.get("location")),
                _safe_excel_value(job.get("department")),
                _safe_excel_value(job.get("id")),
                _safe_excel_value(job.get("apply_url")),
            ])

    # Optional: auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filename)
    log.info(f"\n📊 Saved Excel file: {filename}")


def load_jobs_from_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        job = dict(zip(headers, row))
        log.info(job)
        jobs.append(job)

    return jobs


def _safe_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return value


async def _perform_search(page: Page, search: SearchConfig):
    try:
        box = page.locator(search.selector).first

        if await box.count() == 0:
            log.warning("Search box not found")
            return

        log.info("Performing search...")

        await box.click()
        await box.fill("")  # clear

        await box.type(search.query, delay=search.type_delay_ms)

        if search.press_enter:
            await box.press("Enter")

        await page.wait_for_timeout(search.wait_after_ms)

    except Exception as e:
        log.error(f"Search failed: {e}")


def _extract_uber_location(raw):
    # 1. Prefer allLocations
    all_locs = raw.get("allLocations")

    if isinstance(all_locs, list) and all_locs:
        locations = []
        for loc in all_locs:
            city = loc.get("city")
            region = loc.get("region")
            country = loc.get("countryName")

            parts = [p for p in [city, region, country] if p]
            if parts:
                locations.append(", ".join(parts))

        return " | ".join(locations)

    # 2. fallback to single location
    loc = raw.get("location")
    if isinstance(loc, dict):
        parts = [
            loc.get("city"),
            loc.get("region"),
            loc.get("countryName")
        ]
        return ", ".join([p for p in parts if p])

    return None


async def _paginate_page_number(page: Page, cfg: PaginationConfig, on_page: Callable):
    """
    Click numbered page links and call on_page() after each one.
    on_page is an async callable that scrapes the current page state.
    """
    current = 1
    while current <= cfg.max_pages:
        log.info(f"  → Page {current}")
        await page.wait_for_timeout(cfg.page_load_wait_ms)
        # scroll to load lazy content
        for _ in range(cfg.scroll_rounds):
            await page.mouse.wheel(0, cfg.scroll_step_px)
            await page.wait_for_timeout(cfg.scroll_pause_ms)
        await on_page()
        next_num = current + 1
        sel = cfg.page_number_selector_template.replace("{page}", str(next_num))
        btn = page.locator(sel).first
        if await btn.count() == 0 or not await btn.is_visible():
            log.info("  → No more pages.")
            break
        await btn.scroll_into_view_if_needed()
        await btn.click()
        current += 1


async def _paginate_scroll(page: Page, cfg: PaginationConfig):
    """Scroll down repeatedly to trigger lazy-loading."""
    for _ in range(cfg.scroll_rounds):
        await page.mouse.wheel(0, cfg.scroll_step_px)
        await page.wait_for_timeout(cfg.scroll_pause_ms)


async def _paginate_click(page: Page, cfg: PaginationConfig):
    """Click show-more / load-more buttons up to max_clicks times."""
    count = 0
    while count < cfg.max_clicks:
        clicked = False
        for sel in cfg.click_selectors:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible():
                    log.info(f"  → clicking: {sel}")
                    await btn.scroll_into_view_if_needed()
                    await btn.click()
                    await page.wait_for_timeout(2000)
                    clicked = True
                    count += 1
                    break
            except Exception:
                continue
        if not clicked:
            break
